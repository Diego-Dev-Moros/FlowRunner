# modules/funciones/acciones/io_write.py
from __future__ import annotations
from typing import Any, Dict, Optional

try:
    from modules.funciones.archivos import escritura as mod_escritura
except Exception:
    mod_escritura = None

def escribir_csv(variable: str, ruta: str, contexto: Dict[str, Any]):
    data = contexto.get(variable)
    if mod_escritura and hasattr(mod_escritura, "escribir_csv"):
        mod_escritura.escribir_csv(data, ruta)
    else:
        import pandas as pd
        df = data if isinstance(data, pd.DataFrame) else pd.DataFrame(data)
        df.to_csv(ruta, index=False)
    return {"status": "OK", "message": f"CSV escrito: {ruta}", "preview": ruta}

def escribir_excel(variable: str, ruta: str, contexto: Dict[str, Any],
                   hoja: str = "Sheet1", modo: str = "sobrescribir",
                   inicio_celda: Optional[str] = None, incluir_cabeceras: str = "sí"):
    data = contexto.get(variable)
    if mod_escritura and hasattr(mod_escritura, "escribir_excel"):
        try:
            mod_escritura.escribir_excel(
                data, ruta, hoja=hoja, modo=modo,
                inicio_celda=inicio_celda, incluir_cabeceras=incluir_cabeceras
            )
        except TypeError:
            mod_escritura.escribir_excel(data, ruta)
    else:
        import pandas as pd
        with pd.ExcelWriter(ruta, engine="openpyxl", mode="w") as w:
            (data if isinstance(data, pd.DataFrame) else pd.DataFrame(data)).to_excel(
                w, sheet_name=hoja, index=False
            )
    return {"status": "OK", "message": f"Excel escrito: {ruta}", "preview": ruta}

def excel_crear_hoja(ruta: str, nombre_hoja: str, si_existe: str = "reemplazar"):
    if mod_escritura and hasattr(mod_escritura, "excel_crear_hoja"):
        out = mod_escritura.excel_crear_hoja(ruta, nombre_hoja, si_existe)
        return {"status": "OK", "message": f"Hoja creada: {out}", "preview": out}
    # Fallback mínimo
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    import os
    if os.path.exists(ruta):
        wb = load_workbook(ruta)
    else:
        wb = Workbook()
    if nombre_hoja in wb.sheetnames:
        if si_existe == "reemplazar":
            ws = wb[nombre_hoja]; wb.remove(ws); wb.create_sheet(nombre_hoja)
        elif si_existe == "renombrar":
            i = 2; base = nombre_hoja
            while f"{base}_{i}" in wb.sheetnames: i += 1
            nombre_hoja = f"{base}_{i}"; wb.create_sheet(nombre_hoja)
        else:
            raise ValueError(f"La hoja '{nombre_hoja}' ya existe")
    else:
        wb.create_sheet(nombre_hoja)
    wb.save(ruta)
    return {"status": "OK", "message": f"Hoja creada: {nombre_hoja}", "preview": nombre_hoja}
