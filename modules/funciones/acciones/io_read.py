# modules/funciones/acciones/io_read.py
from __future__ import annotations
from typing import Any, Dict, Optional

try:
    from modules.funciones.archivos import lectura as mod_lectura
except Exception:
    mod_lectura = None

def _preview_table(df, max_rows=10) -> str:
    try:
        import pandas as pd  # noqa
        if df is None: return ""
        return df.head(max_rows).to_string(index=False)
    except Exception:
        return str(df)

def _auto_name(contexto: Dict[str, Any], base: str) -> str:
    cnts = contexto.setdefault("_contadores", {})
    cnts[base] = int(cnts.get(base, 0)) + 1
    return f"{base}_{cnts[base]}"

def leer_csv(ruta: str, contexto: Dict[str, Any], nombre_personalizado: Optional[str] = None):
    if mod_lectura and hasattr(mod_lectura, "leer_csv"):
        df = mod_lectura.leer_csv(ruta)
    else:
        import pandas as pd
        df = pd.read_csv(ruta)
    var = nombre_personalizado or _auto_name(contexto, "csv")
    contexto[var] = df
    contexto["last_df"] = df
    return {"status": "OK", "message": f"CSV leído → {var} ({len(df)} filas)", "preview": _preview_table(df)}

def leer_excel(ruta: str, hoja: str, contexto: Dict[str, Any], nombre_personalizado: Optional[str] = None):
    if mod_lectura and hasattr(mod_lectura, "leer_excel"):
        df = mod_lectura.leer_excel(ruta, hoja)
    else:
        import pandas as pd
        df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
    var = nombre_personalizado or _auto_name(contexto, "excel")
    contexto[var] = df
    contexto["last_df"] = df
    return {"status": "OK", "message": f"Excel leído → {var} ({len(df)} filas)", "preview": _preview_table(df)}

def leer_txt(ruta: str, delimitador: str, contexto: Dict[str, Any], nombre_personalizado: Optional[str] = None):
    if mod_lectura and hasattr(mod_lectura, "leer_txt"):
        df = mod_lectura.leer_txt(ruta, delimitador)  # si tu módulo usa leer_txt_delimitado, ajústalo aquí
    elif mod_lectura and hasattr(mod_lectura, "leer_txt_delimitado"):
        df = mod_lectura.leer_txt_delimitado(ruta, delimitador)
    else:
        import pandas as pd
        df = pd.read_csv(ruta, sep=delimitador)
    var = nombre_personalizado or _auto_name(contexto, "txt")
    contexto[var] = df
    contexto["last_df"] = df
    return {"status": "OK", "message": f"TXT leído → {var} ({len(df)} filas)", "preview": _preview_table(df)}

def excel_leer_rango(ruta: str, hoja: str, contexto: Dict[str, Any],
                     rango: Optional[str] = None, columnas: Optional[str] = None,
                     nombre_personalizado: Optional[str] = None):
    if mod_lectura and hasattr(mod_lectura, "excel_leer_rango"):
        df = mod_lectura.excel_leer_rango(ruta, hoja, rango, columnas)
    else:
        import pandas as pd
        if rango:
            from openpyxl import load_workbook
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb[hoja]
            cells = ws[rango]
            data = [[c.value for c in row] for row in cells]
            wb.close()
            if data and all(isinstance(x, (str, type(None))) for x in data[0]):
                df = pd.DataFrame(data[1:], columns=[str(x) if x is not None else "" for x in data[0]])
            else:
                df = pd.DataFrame(data)
        else:
            df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
            if columnas:
                cols = [c.strip() for c in columnas.split(",") if c.strip()]
                df = df[cols]
    var = nombre_personalizado or _auto_name(contexto, "excel")
    contexto[var] = df
    contexto["last_df"] = df
    return {"status": "OK", "message": f"Excel (parcial) → {var} ({len(df)} filas)", "preview": _preview_table(df)}
