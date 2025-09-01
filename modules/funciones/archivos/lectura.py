import os
import pandas as pd

def leer_csv(ruta: str, **kwargs):
    if not os.path.exists(ruta):
        raise FileNotFoundError(ruta)
    df = pd.read_csv(ruta, **{k:v for k,v in kwargs.items() if v not in (None, "")})
    return df

def leer_excel(ruta: str, hoja: str, **kwargs):
    if not os.path.exists(ruta):
        raise FileNotFoundError(ruta)
    df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
    return df

def leer_txt(ruta: str, delimitador: str = ",", **kwargs):
    if not os.path.exists(ruta):
        raise FileNotFoundError(ruta)
    df = pd.read_csv(ruta, sep=delimitador)
    return df

# NUEVO: Excel parcial
def excel_leer_rango(ruta: str, hoja: str, rango: str | None = None, columnas: str | None = None):
    """Lee rango A1:D100 o columnas por nombre (coma separada)."""
    if not os.path.exists(ruta):
        raise FileNotFoundError(ruta)

    if rango:
        from openpyxl import load_workbook
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb[hoja]
        cells = ws[rango]
        data = [[c.value for c in row] for row in cells]
        wb.close()
        # cabecera si la primera fila parece texto
        if data and all(isinstance(x, (str, type(None))) for x in data[0]):
            return pd.DataFrame(data[1:], columns=[str(x) if x is not None else "" for x in data[0]])
        return pd.DataFrame(data)

    df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
    if columnas:
        cols = [c.strip() for c in columnas.split(",") if c.strip()]
        df = df[cols]
    return df
