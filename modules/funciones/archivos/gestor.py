import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.workbook import Workbook

def escribir_csv(variable, ruta: str, **kwargs):
    _to_df(variable).to_csv(ruta, index=False, **kwargs)
    return ruta

def escribir_excel(variable, ruta: str, hoja: str = "Sheet1", modo: str = "sobrescribir",
                   inicio_celda: str | None = None, incluir_cabeceras: str = "sí"):
    """modo: sobrescribir | append"""
    df = _to_df(variable)
    include_header = (str(incluir_cabeceras).lower() in ("si","sí","yes","true","1"))

    if modo == "append" and os.path.exists(ruta):
        with pd.ExcelWriter(ruta, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            try:
                sh = writer.book[hoja] if hoja in writer.book.sheetnames else None
                startrow = (sh.max_row if sh and sh.max_row else 0)
            except Exception:
                startrow = 0
            df.to_excel(writer, sheet_name=hoja, startrow=startrow, index=False, header=include_header)
        return ruta

    with pd.ExcelWriter(ruta, engine="openpyxl", mode="w") as writer:
        if inicio_celda:
            col_letters, row_num = _split_cell(inicio_celda)
            startcol = column_index_from_string(col_letters) - 1
            df.to_excel(writer, sheet_name=hoja, startrow=row_num - 1, startcol=startcol, index=False, header=include_header)
        else:
            df.to_excel(writer, sheet_name=hoja, index=False, header=include_header)
    return ruta

def excel_crear_hoja(ruta: str, nombre_hoja: str, si_existe: str = "reemplazar"):
    if os.path.exists(ruta):
        wb = load_workbook(ruta)
    else:
        wb = Workbook()

    if nombre_hoja in wb.sheetnames:
        if si_existe == "reemplazar":
            ws = wb[nombre_hoja]; wb.remove(ws); wb.create_sheet(nombre_hoja)
        elif si_existe == "renombrar":
            i = 2; base = nombre_hoja
            while f"{base}_{i}" in wb.sheetnames: i += 1
            nombre_hoja = f"{base}_{i}"; wb.create_sheet(nombre_hoja)
        else:
            raise ValueError(f"La hoja '{nombre_hoja}' ya existe")
    else:
        wb.create_sheet(nombre_hoja)

    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            ws0 = wb['Sheet']
            if ws0.max_row == 1 and ws0.max_column == 1 and ws0['A1'].value is None:
                wb.remove(ws0)
        except Exception:
            pass

    wb.save(ruta)
    return nombre_hoja

# helpers
def _split_cell(cell: str):
    col, row = coordinate_from_string(cell.upper())
    return col, row

def _to_df(value):
    if isinstance(value, pd.DataFrame):
        return value
    if isinstance(value, (list, tuple)):
        return pd.DataFrame(value)
    return pd.DataFrame([value])
