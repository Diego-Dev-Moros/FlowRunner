# modules/utils/data_io.py
"""
Utilidades consolidadas para entrada/salida de datos.
"""
import os
import pandas as pd
from typing import Any, Dict, Optional, List
import glob


def leer_csv(ruta: str, encoding: str = "utf-8", **kwargs) -> pd.DataFrame:
    """Lee un archivo CSV."""
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró el archivo: {ruta}")
    
    try:
        df = pd.read_csv(ruta, encoding=encoding, **{k:v for k,v in kwargs.items() if v not in (None, "")})
        return df
    except UnicodeDecodeError:
        # Fallback a latin-1 si utf-8 falla
        df = pd.read_csv(ruta, encoding='latin-1', **{k:v for k,v in kwargs.items() if v not in (None, "")})
        return df


def leer_excel(ruta: str, hoja: str, **kwargs) -> pd.DataFrame:
    """Lee un archivo Excel."""
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró el archivo: {ruta}")
    
    df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl", **kwargs)
    return df


def excel_leer_rango(ruta: str, hoja: str, rango: str) -> pd.DataFrame:
    """Lee un rango específico de Excel (ej: A1:D100)."""
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró el archivo: {ruta}")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb[hoja]
        cells = ws[rango]
        
        # Extraer datos
        data = []
        for row in cells:
            if isinstance(row, tuple):  # Multiple cells
                data.append([cell.value for cell in row])
            else:  # Single cell
                data.append([row.value])
        
        wb.close()
        
        # Detectar si primera fila son headers
        if data and len(data) > 1:
            first_row = data[0]
            if all(isinstance(x, (str, type(None))) for x in first_row if x is not None):
                headers = [str(x) if x is not None else f"Col_{i}" for i, x in enumerate(first_row)]
                return pd.DataFrame(data[1:], columns=headers)
        
        return pd.DataFrame(data)
        
    except ImportError:
        # Fallback si no hay openpyxl
        df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
        return df


def escribir_csv(df: pd.DataFrame, ruta_destino: str, encoding: str = "utf-8", **kwargs) -> None:
    """Escribe DataFrame a CSV."""
    # Crear directorio si no existe
    os.makedirs(os.path.dirname(ruta_destino), exist_ok=True)
    df.to_csv(ruta_destino, encoding=encoding, index=False, **kwargs)


def escribir_excel(df: pd.DataFrame, ruta_destino: str, hoja: str = "Hoja1", **kwargs) -> None:
    """Escribe DataFrame a Excel."""
    # Crear directorio si no existe
    os.makedirs(os.path.dirname(ruta_destino), exist_ok=True)
    df.to_excel(ruta_destino, sheet_name=hoja, index=False, engine="openpyxl", **kwargs)


def carpeta_listar(ruta: str, patron: str = "*") -> List[str]:
    """Lista archivos en una carpeta con patrón opcional."""
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró la carpeta: {ruta}")
    
    if not os.path.isdir(ruta):
        raise NotADirectoryError(f"La ruta no es una carpeta: {ruta}")
    
    patron_completo = os.path.join(ruta, patron)
    archivos = glob.glob(patron_completo)
    
    # Devolver solo los nombres de archivo, no las rutas completas
    return [os.path.basename(archivo) for archivo in archivos if os.path.isfile(archivo)]


def crear_carpeta(ruta: str) -> None:
    """Crea una carpeta."""
    os.makedirs(ruta, exist_ok=True)


def mover_archivo(origen: str, destino: str, si_existe: str = "sobrescribir") -> None:
    """Mueve un archivo."""
    import shutil
    
    if not os.path.exists(origen):
        raise FileNotFoundError(f"Archivo origen no encontrado: {origen}")
    
    # Crear directorio destino si no existe
    os.makedirs(os.path.dirname(destino), exist_ok=True)
    
    if os.path.exists(destino) and si_existe == "saltar":
        return
    elif os.path.exists(destino) and si_existe == "renombrar":
        base, ext = os.path.splitext(destino)
        contador = 1
        while os.path.exists(f"{base}_{contador}{ext}"):
            contador += 1
        destino = f"{base}_{contador}{ext}"
    
    shutil.move(origen, destino)


def copiar_archivo(origen: str, destino: str, si_existe: str = "sobrescribir") -> None:
    """Copia un archivo."""
    import shutil
    
    if not os.path.exists(origen):
        raise FileNotFoundError(f"Archivo origen no encontrado: {origen}")
    
    # Crear directorio destino si no existe
    os.makedirs(os.path.dirname(destino), exist_ok=True)
    
    if os.path.exists(destino) and si_existe == "saltar":
        return
    elif os.path.exists(destino) and si_existe == "renombrar":
        base, ext = os.path.splitext(destino)
        contador = 1
        while os.path.exists(f"{base}_{contador}{ext}"):
            contador += 1
        destino = f"{base}_{contador}{ext}"
    
    shutil.copy2(origen, destino)


def eliminar_archivo(ruta: str) -> None:
    """Elimina un archivo."""
    if os.path.exists(ruta):
        os.remove(ruta)
